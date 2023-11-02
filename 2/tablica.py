# skripta za generiranje py dicta koji predstavlja tablicu
import openpyxl

workbook = openpyxl.load_workbook('tablica.xlsx')

sheet = workbook.active

with open('output.txt', 'w') as text_file:
    for row in sheet.iter_rows(min_row=4, min_col=3, max_row=20, max_col=16):
        for cell in row:
            if cell.value != "Odbaci":
                column_index = cell.column
                row_index = 3
                value_in_third_row = sheet.cell(row=row_index, column=column_index).value

                column_index = 2
                row_index = cell.row
                value_in_second_column = sheet.cell(row=row_index, column=column_index).value

                text_file.write(f'("{value_in_second_column}", "{value_in_third_row}"): {cell.value},\n')